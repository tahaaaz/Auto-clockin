from flask import Flask, render_template_string, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
FILE_NAME = "leave_data.xlsx"

# Create file with headers if it doesn't exist
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Leave Date", "Reason", "Status"])
    wb.save(FILE_NAME)

# HTML Form template
form_template = """
<!DOCTYPE html>
<html>
<head>
    <title>Leave Application</title>
</head>
<body>
    <h2>Leave Application Form</h2>
    <form method="post">
        <label>Name:</label><br>
        <input type="text" name="name" required><br><br>

        <label>Leave Date:</label><br>
        <input type="date" name="leave_date" required><br><br>

        <label>Reason:</label><br>
        <input type="text" name="reason" required><br><br>

        <label>Status:</label><br>
        <select name="status">
            <option value="Pending">Pending</option>
            <option value="Approved">Approved</option>
            <option value="Rejected">Rejected</option>
        </select><br><br>

        <button type="submit">Submit</button>
    </form>
</body>
</html>
"""

@app.route("/", methods=["GET", "POST"])
def leave_form():
    if request.method == "POST":
        name = request.form["name"]
        leave_date = request.form["leave_date"]
        reason = request.form["reason"]
        status = request.form["status"]

        wb = load_workbook(FILE_NAME)
        ws = wb.active
        ws.append([name, leave_date, reason, status])
        wb.save(FILE_NAME)

        return redirect("/")  # refresh after submit

    return render_template_string(form_template)

if __name__ == "__main__":
    app.run(debug=True)
